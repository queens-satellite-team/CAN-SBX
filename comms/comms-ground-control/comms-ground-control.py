"""
authur: Sean Tedesco
email:  17sart@queensu.ca
Date: July 20th, 2021 

1. Make sure to install all the required packages with pip! 
$ pip install -r requirements.txt

2. Update any parameters in the "user config" section! 

- please feel free to reach out for any and all questions regarding this wee bit of code - 
"""
###########################################################################################
######################################### imports #########################################
import serial
import openpyxl as pxl 
import time

###########################################################################################
####################################### user config #######################################
port_name = "/dev/cu.usbserial-1420"
workbook_name = "can-sbx-launch-data"

###########################################################################################
######################################### globals #########################################
start_marker = "<"
end_marker = ">"
data_started = False
data_buf = ""
message_complete = False

time_column = "A"
flag_column = "B"
message_column = "C"

###########################################################################################
##################################### sub-routines ########################################

def setup_serial(baud_rate, serial_port_name):

    global serialPort

    try: 
        serialPort = serial.Serial(port=serial_port_name, baudrate=baud_rate, timeout=0, rtscts=True)
        print("Serial port " + serial_port_name + " opened  Baudrate " + str(baud_rate))
        wait_for_arduino()
    except serial.SerialException as e:
        print(e)
        print("<is the Arduino connected?>")

def wait_for_arduino():

    print("waiting for arduino to reset")

    msg = ""
    while msg.find("success: arduino is ready") == -1:
        msg = recv_from_arduino()
        if not (msg == "XXX"):
            print(msg)

def recv_from_arduino():

    global start_marker, end_marker, serialPort, data_started, data_buf, message_complete

    if serialPort.inWaiting() > 0 and message_complete == False:
        x = serialPort.read().decode("utf-8")

        if data_started == True:
            if x != end_marker:
                data_buf = data_buf + x
            else:
                data_started = False
                message_complete = True
        elif x == start_marker:
            data_buf = ""
            data_started = True

    if message_complete == True:
        message_complete = False
        return data_buf
    else:
        return "XXX"

def init_workbook(name_of_workbook: str="my_workbook"):
    global workbook, path, current_row, current_column

    path = "comms/comms-ground-control/"
    workbook = pxl.Workbook()
    current_row = 1

    if is_workbook(name_of_workbook):
        print("<" + name_of_workbook + " already made>")
    else: 
        save_workbook(name_of_workbook)
        print("<created " + name_of_workbook + " workbook>")
        
def is_workbook(name_of_workbook: str="my_workbook"):
    global path 

    file_name = path + name_of_workbook + ".xlsx"
    
    try:
        pxl.load_workbook(file_name)
    except FileNotFoundError  as e:
        print(e)
        return False 
    return True

def save_workbook(name_of_workbook: str="my_workbook"):
    global workbook, path

    filename = path + name_of_workbook + '.xlsx'
    print(f"<saving {filename}... ", end='')
    try:
        workbook.save(filename)
        print("done>")
    except:
        print("could not save>")

def set_sheet(name_of_sheet: str="my_sheet", location_in_workbook: int=0):
    global workbook
    return workbook.create_sheet(name_of_sheet, location_in_workbook)

def get_sheet_by_name(name_of_sheet:str="my_sheet"):
    global workbook
    return workbook[name_of_sheet]

def remove_sheet_by_name(name_of_sheet:str="my_sheet"):
    global workbook
    workbook.remove(name_of_sheet)

def _set_data_in_sheet(sheet, data, row:str="1", column:str="A"):
    sheet[column+row] = data

def get_data_in_sheet(sheet, row:str="1", column:str="A"):
    return sheet[column+row]

def _set_time_in_sheet(sheet, time):
    global current_row, time_column
    _set_data_in_sheet(sheet, time, str(current_row), time_column)

def _set_msg_in_sheet(sheet, msg):
    global current_row, message_column
    _set_data_in_sheet(sheet, msg, str(current_row), message_column)

def _set_flag_in_sheet(sheet, flag):
    global current_row, flag_column
    _set_data_in_sheet(sheet, flag, str(current_row), flag_column)

def set_response_in_sheet(sheet, response):
    global current_row, time_column

    _set_data_in_sheet(sheet, response, str(current_row), time_column)
    current_row += 1

###########################################################################################
########################################## main ###########################################
def main():

    setup_serial(115200, port_name)
    init_workbook(workbook_name) 
    telemtry_page = set_sheet("telemtry data")

    while True:
        try:
            arduino_response = recv_from_arduino()
            if not (arduino_response == "XXX"):
                print(f"<time:{time.time()}#response:{arduino_response}")
                set_response_in_sheet(telemtry_page, arduino_response)
        except KeyboardInterrupt as error:
            print(f"closing control due to user interupt: {error}")
            break
        
    save_workbook(workbook_name)

if __name__ == '__main__':
    main()